import openpyxl as xl
from datetime import datetime

banks = {'bank_lbp.xlsx': 'lbp.xlsx', 'bank_usd.xlsx': 'usd.xlsx', 'bank_euro.xlsx': 'euro.xlsx'}
banks = {'bank_usd.xlsx': 'usd.xlsx'}

for key, value in banks.items():

    bank_wb = xl.load_workbook(key)
    bank_sheet = bank_wb.active

    wb = xl.Workbook()
    sheet = wb.active
    myList = []

    for row in bank_sheet.iter_rows(min_row=1, max_row=8, values_only=True):
        sheet.append(row)

    for row in bank_sheet.iter_rows(min_row=9, values_only=True):
        myList.append(row)

    ctr = 9
    for row in reversed(myList):
        sheet.append(row)
        if isinstance(row[1], datetime):
            sheet.cell(row=ctr, column=2).value = row[1].strftime("%m/%d/%Y")
        if isinstance(row[2], datetime):
            sheet.cell(row=ctr, column=3).value = row[2].strftime("%m/%d/%Y")
        ctr += 1

    """
    cell = sheet['B9']
    value = cell.value
    cell.value = value.strftime("%m/%d/%Y")
    """
    wb.save(value)
